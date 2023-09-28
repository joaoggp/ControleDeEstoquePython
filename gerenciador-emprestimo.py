import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Função para cadastrar um novo equipamento
def cadastrar_equipamento(equipamentos, nome, quantidade):
    equipamento_id = len(equipamentos) + 1
    equipamento = {"ID": equipamento_id, "Nome": nome, "Quantidade": quantidade}
    equipamentos.append(equipamento)
    atualizar_planilha_equipamentos(equipamentos)

# Função para listar todos os equipamentos cadastrados
def listar_equipamentos(equipamentos):
    if not equipamentos:
        print("Não há equipamentos cadastrados.")
    else:
        print("Equipamentos cadastrados:")
        for equipamento in equipamentos:
            print(f"ID: {equipamento['ID']} | Nome: {equipamento['Nome']} | Quantidade: {equipamento['Quantidade']}")

# Função para atualizar a planilha de equipamentos
def atualizar_planilha_equipamentos(equipamentos):
    workbook = openpyxl.load_workbook('equipamentos.xlsx')
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)  # Limpa os dados antigos

    for equipamento in equipamentos:
        sheet.append([equipamento['ID'], equipamento['Nome'], equipamento['Quantidade']])

    workbook.save('equipamentos.xlsx')

# Função para listar empréstimos ativos com ID diferente de None
def listar_emprestimos_ativos(emprestimos):
    emprestimos_ativos = [emprestimo for emprestimo in emprestimos if emprestimo['ID'] is not None]

    if not emprestimos_ativos:
        print("Não há empréstimos ativos com ID.")
    else:
        print("Empréstimos ativos:")
        for emprestimo in emprestimos_ativos:
            print(f"ID: {emprestimo['ID']} | Usuário: {emprestimo['Usuário']} | Matrícula: {emprestimo['Matrícula']} | Equipamentos: {', '.join(map(str, emprestimo['Equipamentos']))} | Data Início: {emprestimo['Data Início']} | Data Fim: {emprestimo['Data Fim']}")

# Função para limpar empréstimos com ID igual a None
def limpar_emprestimos_nulos(emprestimos):
    emprestimos[:] = [emprestimo for emprestimo in emprestimos if emprestimo['ID'] is not None]

# Função para verificar conflitos entre empréstimos ativos
def verificar_conflitos(emprestimos, novo_emprestimo):
    for emprestimo in emprestimos:
        if emprestimo['Data Início'] <= novo_emprestimo['Data Fim'] and emprestimo['Data Fim'] >= novo_emprestimo['Data Início']:
            conflitantes = set(emprestimo['Equipamentos']).intersection(set(novo_emprestimo['Equipamentos']))
            if conflitantes:
                return f"Conflito com empréstimo ID {emprestimo['ID']}. Equipamentos em conflito: {', '.join(map(str, conflitantes))}"
    return None

# Função para realizar um empréstimo
def realizar_emprestimo(emprestimos, equipamentos, usuario, matricula, equipamento_ids, data_inicio, data_fim):
    emprestimo_id = len(emprestimos) + 1
    emprestimo = {"ID": emprestimo_id, "Usuário": usuario, "Matrícula": matricula, "Equipamentos": equipamento_ids, "Data Início": data_inicio, "Data Fim": data_fim}
    emprestimos.append(emprestimo)
    
    # Reduz a quantidade disponível de equipamentos emprestados
    for equipamento in equipamentos:
        if equipamento['ID'] in equipamento_ids:
            quantidade_emprestada = equipamento_ids.count(equipamento['ID'])
            equipamento['Quantidade'] -= quantidade_emprestada
    
    atualizar_planilha_equipamentos(equipamentos)
    atualizar_planilha_emprestimos(emprestimos)

    print("Empréstimo realizado com sucesso!\n")
    print("Empréstimos Ativos:")
    listar_emprestimos_ativos(emprestimos)

# Função para atualizar a planilha de empréstimos ativos
def atualizar_planilha_emprestimos(emprestimos):
    workbook = openpyxl.load_workbook('emprestimos-ativos.xlsx')
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)  # Limpa os dados antigos

    for emprestimo in emprestimos:
                sheet.append([emprestimo['ID'], emprestimo['Usuário'], emprestimo['Matrícula'], ', '.join(map(str, emprestimo['Equipamentos'])), emprestimo['Data Início'], emprestimo['Data Fim']])

    workbook.save('emprestimos-ativos.xlsx')

# Exemplo de uso
if __name__ == "__main__":
    equipamentos = []
    emprestimos = []

    # Carrega os equipamentos existentes do arquivo Excel
    try:
        workbook_equipamentos = openpyxl.load_workbook('equipamentos.xlsx')
        sheet_equipamentos = workbook_equipamentos.active
        for row in sheet_equipamentos.iter_rows(min_row=2, values_only=True):
            equipamento_id = row[0]
            nome_equipamento = row[1]
            quantidade_equipamento = row[2]
            equipamento = {"ID": equipamento_id, "Nome": nome_equipamento, "Quantidade": quantidade_equipamento}
            equipamentos.append(equipamento)
    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo
        workbook_equipamentos = Workbook()
        sheet_equipamentos = workbook_equipamentos.active
        sheet_equipamentos.append(["ID", "Nome", "Quantidade"])
        workbook_equipamentos.save('equipamentos.xlsx')

    # Carrega os empréstimos ativos existentes do arquivo Excel
    try:
        workbook_emprestimos = openpyxl.load_workbook('emprestimos-ativos.xlsx')
        sheet_emprestimos = workbook_emprestimos.active
        for row in sheet_emprestimos.iter_rows(min_row=2, values_only=True):
            emprestimo_id = row[0]
            usuario = row[1]
            matricula = row[2]
            equipamentos_str = row[3]
            equipamentos = [int(x) for x in equipamentos_str.split(', ')] if equipamentos_str else []
            data_inicio = row[4]
            data_fim = row[5]
            emprestimo = {"ID": emprestimo_id, "Usuário": usuario, "Matrícula": matricula, "Equipamentos": equipamentos, "Data Início": data_inicio, "Data Fim": data_fim}
            emprestimos.append(emprestimo)

        limpar_emprestimos_nulos(emprestimos)  # Limpa empréstimos com ID None
    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo
        workbook_emprestimos = Workbook()
        sheet_emprestimos = workbook_emprestimos.active
        sheet_emprestimos.append(["ID", "Usuário", "Matrícula", "Equipamentos", "Data Início", "Data Fim"])
        workbook_emprestimos.save('emprestimos-ativos.xlsx')

    # Menu de opções
    while True:
        print("\nMENU:")
        print("1 - Cadastrar Equipamento")
        print("2 - Listar Equipamentos")
        print("3 - Realizar Empréstimo")
        print("4 - Listar Empréstimos Ativos")
        print("5 - Verificar Conflitos de Empréstimos")
        print("6 - Sair")

        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            nome_equipamento = input("Digite o nome do equipamento: ")
            quantidade_equipamento = int(input("Digite a quantidade do equipamento: "))
            cadastrar_equipamento(equipamentos, nome_equipamento, quantidade_equipamento)
            print("Equipamento cadastrado com sucesso!")

        elif opcao == "2":
            listar_equipamentos(equipamentos)

        elif opcao == "3":
            listar_equipamentos(equipamentos)
            usuario = input("Digite o nome do usuário responsável: ")
            matricula = input("Digite a matrícula do usuário: ")
            equipamento_ids = [int(x) for x in input("Digite os IDs dos equipamentos separados por vírgula: ").split(',')]
            data_inicio = input("Digite a data de início no formato DD-MM-YYYY: ")
            data_fim = input("Digite a data de fim no formato DD-MM-YYYY: ")
            novo_empréstimo = {"Usuário": usuario, "Matrícula": matricula, "Equipamentos": equipamento_ids, "Data Início": data_inicio, "Data Fim": data_fim}
            conflito = verificar_conflitos(emprestimos, novo_empréstimo)
            if conflito:
                print(f"Conflito encontrado: {conflito}")
            else:
                realizar_emprestimo(emprestimos, equipamentos, usuario, matricula, equipamento_ids, data_inicio, data_fim)

        elif opcao == "4":
            listar_emprestimos_ativos(emprestimos)

        elif opcao == "5":
            usuario = input("Digite o nome do usuário para verificar conflitos: ")
            matricula = input("Digite a matrícula do usuário: ")
            data_inicio = input("Digite a data de início do empréstimo no formato DD-MM-YYYY: ")
            data_fim = input("Digite a data de fim do empréstimo no formato DD-MM-YYYY: ")
            equipamento_ids = [int(x) for x in input("Digite os IDs dos equipamentos separados por vírgula: ").split(', ')]
            novo_empréstimo = {"Usuário": usuario, "Matrícula": matricula, "Equipamentos": equipamento_ids, "Data Início": data_inicio, "Data Fim": data_fim}
            conflito = verificar_conflitos(emprestimos, novo_empréstimo)
            if conflito:
                print(f"Conflito encontrado: {conflito}")
            else:
                print("Nenhum conflito encontrado.")

        elif opcao == "6":
            print("Encerrando o programa.")
            break
        else:
            print("Opção inválida. Tente novamente.")

